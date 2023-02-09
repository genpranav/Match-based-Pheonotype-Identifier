import openpyxl

filename = input('enter the file name(dataset name) :')
wb = openpyxl.load_workbook(filename)
sh1 = wb['Sheet1']

#print(wb.active.title)

rowmax =  sh1.max_row
columnmax = sh1.max_column

phenotype = []
geneinfo = []

sdnawidth = columnmax - 2

for i in range(2,rowmax+1):
    phenotype.append(sh1.cell(i,2).value)
    #print(phenotype[:])

    for j in range(3,columnmax+1):
        geneinfo.append(sh1.cell(i,j).value)
#print(geneinfo)

comp = []
temp = []
celements = []
ratiop = []
ration = []
resultingratio = []
comptracker = []

scanwidth = input('enter the number of individual segments we have to compare :')
scansize = int(scanwidth)

for y in range(2,sdnawidth-scansize+3):
    for g in range(scansize,0,-1):
        j = y-g+scansize-2
        celements.append(j)

    for ind in celements:
        comp.append(geneinfo[ind])
        compst = ''.join(comp)
        compstf = compst.upper()

    #print(comp)
    print(compstf)
    comptracker.append(compstf)

    counttrue = 0
    countfalse = 0


    for n in range(1,rowmax-1):
        selements = range((sdnawidth*n),sdnawidth*(n+1),1)
        #print(selements)
        for indx in selements:
            temp.append(geneinfo[indx])
        tempst = ''.join(temp)
        tempstf = tempst.upper()
        #print(temp)
        print(tempstf)

        #comparisn code


        eval = tempstf.find(compstf)
        if phenotype[n] == 't':
            if eval>=0 and eval<sdnawidth:
                counttrue += 1
        else:
            if eval >= 0 and eval < sdnawidth:
                countfalse += 1
        temp.clear()
        celements.clear()

    ratiopositive = 100*counttrue/((sdnawidth-1)*(rowmax-2))
    n1 = round(ratiopositive,2)
    rationegetive = 100*countfalse/((sdnawidth-1)*(rowmax-2))
    n2 = round(rationegetive,2)
    ratiop.append(n1)
    ration.append(n2)
    #print(ratiop)
    #print(ration)

    print(f'this is the no of times match in phenotype due to gene segment({compstf}) is obsereved = {counttrue}')
    print(f'this is the no of times where gene segment({compstf}) is observed but does not match the phenotype = {countfalse}')
    comp.clear()

zip_object = zip(ratiop,ration)
for ratiopi,rationi in zip_object:
    resultingratio.append(ratiopi-rationi)
print(resultingratio)

maxval = max(resultingratio)
print(max(resultingratio))
genelocation = resultingratio.index(maxval)
print(resultingratio.index(maxval))

print(comptracker)
print(f'the gene segment most probable to have caused the phenotype is {comptracker[genelocation]}')
#print(comp)